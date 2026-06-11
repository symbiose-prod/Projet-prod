"""
pages/reconciliation_transport.py
=================================
Page Réconciliation transport — rapprochement factures transporteur (Pennylane)
↔ commandes Easy Beer (export Excel).

Flux opérateur :
  1. Déposer l'export Easy Beer (.xlsx, feuille « Commandes ») — il RESTE ensuite
     en place (sauvegardé sur disque par tenant) ; bouton pour l'enlever/remplacer.
  2. Choisir une période (libre) — par défaut le mois précédent.
  3. « Lancer » → télécharge + parse les factures Pennylane de la période,
     rapproche avec les commandes, affiche 3 niveaux :
       a) Synthèse (cartes KPIs)
       b) Tableau décisionnel (colonnes clés)
       c) Tableau général (toutes les colonnes EB + calculées) + export Excel

La logique métier vit dans core/reconciliation/ (cœur sans UI, partagé) :
  charger_sources(export_xlsx, date_min, date_max) -> (factures, commandes)
  reconcilier(factures, commandes)                 -> Resultat(kpis, lignes, …)

⚠️ L'appel réseau (téléchargement des PDF) est BLOQUANT et potentiellement long
sur une grande période : il tourne dans un thread + overlay de chargement.

Formats nombres (écran + Excel) : 1 décimale, virgule décimale, « % » pour les
pourcentages, et 3 décimales pour le €/kg facturé. L'Excel écrit les nombres en
texte formaté à la française (virgule garantie quelle que soit la locale Excel).

Persistance :
  - l'export EB est stocké sur disque (data/reconciliation_uploads/, par tenant) et
    son nom dans app.storage.user → il survit aux navigations ET aux redémarrages.
  - le dernier résultat calculé est gardé en cache mémoire (clé tenant_id+email).
"""
from __future__ import annotations

import asyncio
import io
import logging
import os
from datetime import date, timedelta
from pathlib import Path

from nicegui import app, ui

from core.reconciliation.pennylane_cache import PennylaneCache
from core.reconciliation.stockage import repartition_par_enseigne, synthese_stockage
from pages.auth import require_auth
from pages.theme import (
    COLORS,
    confirm_dialog,
    error_banner,
    kpi_card,
    page_layout,
    section_title,
)

_log = logging.getLogger("ferment.reconciliation_transport")

# Dossier où l'export EB est conservé (gitignoré). Un fichier par tenant.
_UPLOAD_DIR = Path("data/reconciliation_uploads")

# Cache disque des factures Pennylane parsées — PARTAGÉ entre tous les
# utilisateurs (les factures SOFRIPA sont les mêmes pour toute la société).
_CACHE = PennylaneCache()

# Cache mémoire serveur du dernier run, par (tenant_id, email).
_RUN_CACHE: dict[tuple, dict] = {}

# Mapping statut -> couleur Quasar (badge) — partagé tableau décisionnel + méga.
_STATUT_BADGE_JS = r"""
    <q-td :props="props">
      <q-badge :color="{
        'OK':'green-6',
        'Écart notable':'orange-7',
        'À vérifier (négatif)':'red-6',
        'Palette (pas de poids)':'blue-grey-5',
        'Pas de poids EB':'grey-6'
      }[props.value] || 'grey-6'" :label="props.value" />
    </q-td>
"""

# Mapping statut -> (fond pastel, couleur texte) pour l'Excel — teintes douces.
_STATUT_FILL = {
    "OK": ("DCFCE7", "166534"),
    "Écart notable": ("FFEDD5", "9A3412"),
    "À vérifier (négatif)": ("FEE2E2", "991B1B"),
    "Palette (pas de poids)": ("E2E8F0", "334155"),
    "Pas de poids EB": ("F1F5F9", "475569"),
}


# ─── Helpers de format (1 décimale, virgule, % ; €/kg à 3 décimales) ─────────

def _periode_defaut() -> tuple[str, str]:
    """Premier et dernier jour du mois précédent, au format ISO 'YYYY-MM-DD'."""
    today = date.today()
    premier_du_mois = date(today.year, today.month, 1)
    dernier = premier_du_mois - timedelta(days=1)          # dernier jour du mois précédent
    premier = date(dernier.year, dernier.month, 1)
    return premier.isoformat(), dernier.isoformat()


def _eur(v) -> str:
    return f"{v:,.1f} €".replace(",", " ").replace(".", ",") if v is not None else "—"


def _eur_kg(v) -> str:
    """€/kg facturé — 3 décimales."""
    return f"{v:,.3f} €".replace(",", " ").replace(".", ",") if v is not None else "—"


def _kg(v) -> str:
    return f"{v:,.1f} kg".replace(",", " ").replace(".", ",") if v is not None else "—"


def _pct(v) -> str:
    return f"{v * 100:.1f} %".replace(".", ",") if v is not None else "—"


def _fr(v, dec: int):
    """Nombre formaté à la française (virgule), sans unité — pour l'Excel. None -> None."""
    if v is None:
        return None
    return f"{v:,.{dec}f}".replace(",", " ").replace(".", ",")


def _fr_pct(v):
    """Pourcentage français pour l'Excel : fraction -> 'X,X %'."""
    if v is None:
        return None
    return f"{v * 100:.1f}".replace(".", ",") + " %"


# Colonnes calculées : (label, getter_valeur_brute, formateur_écran, formateur_Excel).
# Libellés EXACTS de l'onglet « Réconciliation » de référence, ordre 48→58.
_CALC_COLS: list[tuple] = [
    ("N° OT (facture)", lambda L: L.ot, lambda v: v or "—", lambda v: v),
    ("N° pièce (facture)", lambda L: L.piece, lambda v: v or "—", lambda v: v),
    ("Poids Easy Beer (kg)", lambda L: L.poids_eb, _kg, lambda v: _fr(v, 1)),
    ("Poids SOFRIPA (kg)", lambda L: L.poids_sofripa, _kg, lambda v: _fr(v, 1)),
    ("Écart poids (kg)", lambda L: L.ecart_kg, _kg, lambda v: _fr(v, 1)),
    ("Écart poids (%)", lambda L: L.ecart_pct, _pct, _fr_pct),
    ("Coût transport SOFRIPA (€)", lambda L: L.cout_transport, _eur, lambda v: _fr(v, 1)),
    ("Montant HT Easy Beer (€)", lambda L: L.montant_ht, _eur, lambda v: _fr(v, 1)),
    ("Transport / HT (%)", lambda L: L.transport_sur_ht, _pct, _fr_pct),
    ("€ / kg facturé", lambda L: L.eur_par_kg, _eur_kg, lambda v: _fr(v, 3)),
    ("Statut poids", lambda L: L.statut, lambda v: v or "—", lambda v: v),
]


def _eb_headers(res) -> list[str]:
    """Colonnes EB (dans l'ordre de l'export) extraites de la 1re ligne réconciliée."""
    for L in res.lignes:
        if L.commande and L.commande.brut:
            return list(L.commande.brut.keys())
    return []


def _eb_display(v):
    """Valeur EB : floats avec virgule (entiers sans décimale), reste inchangé.

    Sert à l'écran ET à l'Excel (virgule garantie même en locale Excel anglaise).
    """
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else f"{v}".replace(".", ",")
    if v is None or isinstance(v, (str, int)):
        return v
    return str(v)


def _safe_tenant(tenant_id) -> str:
    s = "".join(c for c in str(tenant_id or "default") if c.isalnum() or c in "-_")
    return s or "default"


def _export_path_for(tenant_id) -> Path:
    return _UPLOAD_DIR / f"eb_export_{_safe_tenant(tenant_id)}.xlsx"


def _build_mega_xlsx(res, eb_headers: list[str]) -> bytes:
    """Méga tableau .xlsx mis en forme : entêtes colorés (EB vert / calc orange),
    écarts colorés (vert ≥0 / rouge <0), statut coloré, nombres à la française
    (virgule), colonnes élargies, 1re colonne + entête figées."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Réconciliation"
    n_eb = len(eb_headers)
    labels = [c[0] for c in _CALC_COLS]
    headers = eb_headers + labels
    ws.append(headers)

    i_ecart_kg = labels.index("Écart poids (kg)")
    i_ecart_pct = labels.index("Écart poids (%)")
    i_statut = labels.index("Statut poids")

    green_font = Font(color="16A34A", bold=True)
    red_font = Font(color="DC2626", bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    r = 2
    for L in res.lignes:
        brut = L.commande.brut if (L.commande and L.commande.brut) else {}
        ci = 1
        for h in eb_headers:
            ws.cell(r, ci, _eb_display(brut.get(h)))
            ci += 1
        for j, (lab, getter, _disp, excel_fmt) in enumerate(_CALC_COLS):
            raw = getter(L)
            cell = ws.cell(r, ci, excel_fmt(raw))
            if j in (i_ecart_kg, i_ecart_pct) and isinstance(raw, (int, float)):
                cell.font = green_font if raw >= 0 else red_font
            elif j == i_statut:
                style = _STATUT_FILL.get(raw)
                if style:
                    fill_hex, font_hex = style
                    cell.fill = PatternFill("solid", fgColor=fill_hex)
                    cell.font = Font(color=font_hex, bold=True)
                    cell.alignment = center
            ci += 1
        r += 1

    # Entêtes : EB vert, colonnes calculées orange (plus intense), texte blanc gras.
    eb_fill = PatternFill("solid", fgColor="15803D")
    calc_fill = PatternFill("solid", fgColor="C2410C")
    head_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(1, ci)
        cell.fill = eb_fill if ci <= n_eb else calc_fill
        cell.font = white_bold
        cell.alignment = head_center

    # Largeurs adaptées au contenu (élargies, pas entassées).
    for ci in range(1, len(headers) + 1):
        maxlen = len(str(headers[ci - 1]))
        for rr in range(2, ws.max_row + 1):
            v = ws.cell(rr, ci).value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 4, 12), 45)

    ws.freeze_panes = "B2"      # fige 1re colonne + entête
    ws.row_dimensions[1].height = 32
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inject_mega_css():
    """CSS du méga tableau : scroll horizontal, 1re colonne + entête figées,
    entêtes colorés (EB vert / colonnes calculées à droite en orange)."""
    ui.add_head_html("""
    <style>
      .mega-table .q-table__middle { max-height: 60vh; }
      .mega-table thead tr th { position: sticky; top: 0; z-index: 1; background: #fff; }
      .mega-table tbody td:first-child { position: sticky; left: 0; background: #fff;
                                         z-index: 2; font-weight: 600; }
      .mega-table thead tr th:first-child { position: sticky; left: 0; z-index: 3; }
      .mega-table thead th.mega-eb-head { background: #15803D !important; color: #fff !important; }
      .mega-table thead th.mega-calc-head { background: #C2410C !important; color: #fff !important; }
      /* KPIs : cartes assez larges (2 rangées si besoin) + valeurs sur UNE ligne */
      .reconcil-kpis .kpi-card { min-width: 215px; }
      .reconcil-kpis .kpi-card .text-h6 { white-space: nowrap; }
    </style>
    """)


# ─── Page ───────────────────────────────────────────────────────────────────

@ui.page("/reconciliation-transport")
def page_reconciliation_transport():
    user = require_auth()
    if not user:
        return

    _inject_mega_css()

    tenant_id = user.get("tenant_id")
    cache_key = (tenant_id, user.get("email"))
    cached = _RUN_CACHE.get(cache_key)

    export_path = _export_path_for(tenant_id)
    file_ok = export_path.exists()

    # État local (mutable via closures).
    state: dict = {
        "export_path": str(export_path) if file_ok else None,
        "export_name": app.storage.user.get("reconcil_export_name") if file_ok else None,
        "date_min": (cached or {}).get("date_min") or app.storage.user.get("reconcil_date_min"),
        "date_max": (cached or {}).get("date_max") or app.storage.user.get("reconcil_date_max"),
    }

    with page_layout("Réconciliation transport", "compare_arrows", "/reconciliation-transport"):
        ui.label(
            "Rapprochement des factures transporteur (Pennylane) avec les commandes "
            "Easy Beer. L'export Easy Beer reste en place une fois déposé."
        ).classes("text-body2").style(f"color: {COLORS['ink2']}")

        # ── Avertissement « grande période = lent » ──────────────────────
        with ui.card().classes("w-full").props("flat bordered").style(
            f"border-color: {COLORS['warning']}55; background: {COLORS['warning']}0D"
        ):
            with ui.card_section().classes("row items-center gap-3 q-pa-sm"):
                ui.icon("schedule", size="sm").style(f"color: {COLORS['warning']}")
                ui.label(
                    "Une grande période télécharge et analyse de nombreux PDF : "
                    "le traitement peut prendre plusieurs minutes."
                ).classes("text-body2").style(f"color: {COLORS['ink2']}")

        # ── Paramètres ────────────────────────────────────────────────────
        with ui.card().classes("w-full").props("flat bordered"):
            with ui.card_section().classes("q-pa-md column gap-3"):
                section_title("Paramètres", "tune")

                # Période (libre) — les commandes viennent de l'API Easy Beer en
                # direct ; la bascule « export Excel » est repliée dans Dépannage.
                ui.label("Période (libre — mois, plusieurs mois, ou plage complète)").classes(
                    "text-subtitle2"
                ).style(f"color: {COLORS['ink']}; font-weight: 600")

                d_def_min, d_def_max = _periode_defaut()
                d_min_iso = state["date_min"] or d_def_min
                d_max_iso = state["date_max"] or d_def_max
                with ui.row().classes("w-full gap-4 items-end"):
                    with ui.column().classes("gap-1"):
                        ui.label("Du").classes("text-caption").style(
                            f"color: {COLORS['ink2']}"
                        )
                        date_min = ui.input(value=d_min_iso).props("outlined dense type=date")
                    with ui.column().classes("gap-1"):
                        ui.label("Au").classes("text-caption").style(
                            f"color: {COLORS['ink2']}"
                        )
                        date_max = ui.input(value=d_max_iso).props("outlined dense type=date")
                    # Relance manuelle (discrète) — la réconciliation se lance déjà
                    # automatiquement à l'arrivée et à chaque changement.
                    run_btn = ui.button(
                        "Relancer", icon="refresh",
                    ).props("outline dense color=green-8")

                def _update_run_state():
                    """Source API : pas de fichier requis. Source export : fichier obligatoire."""
                    need_file = source_radio.value == "export"
                    has_file = bool(state.get("export_path")) and os.path.exists(
                        state.get("export_path") or ""
                    )
                    if need_file and not has_file:
                        run_btn.disable()
                    else:
                        run_btn.enable()

                # ── Dépannage (replié) : source des commandes + cache ─────────
                default_source = app.storage.user.get("reconcil_source", "api")
                with ui.expansion(
                    "Dépannage — source des commandes & cache",
                    icon="build",
                    value=(default_source == "export"),
                ).classes("w-full q-mt-sm").props("dense"):
                    ui.label(
                        "Par défaut, les commandes sont récupérées en direct via "
                        "l'API Easy Beer. En cas de problème (panne API…), basculez "
                        "ici sur un export Excel déposé à la main."
                    ).classes("text-caption").style(f"color: {COLORS['ink2']}")

                    source_radio = ui.radio(
                        {"api": "API Easy Beer (direct)", "export": "Export uploadé (secours)"},
                        value=default_source,
                    ).props("inline dense")

                    export_box = ui.column().classes("w-full gap-2")
                    export_box.visible = default_source == "export"

                    def _on_source_change():
                        src = source_radio.value
                        app.storage.user["reconcil_source"] = src
                        export_box.visible = src == "export"
                        _update_run_state()
                        _schedule_run()

                    source_radio.on_value_change(_on_source_change)

                    with export_box, ui.row().classes("items-center gap-2"):
                        if file_ok:
                            _txt = f"✓ {state['export_name'] or 'export'} (conservé)."
                            _col = COLORS["success"]
                        else:
                            _txt = "Aucun fichier déposé."
                            _col = COLORS["ink2"]
                        upload_status = ui.label(_txt).classes("text-caption").style(
                            f"color: {_col}"
                        )

                        def _remove_file():
                            try:
                                if export_path.exists():
                                    export_path.unlink()
                            except OSError:
                                _log.exception("Suppression export EB échouée")
                            app.storage.user.pop("reconcil_export_name", None)
                            state["export_path"] = None
                            state["export_name"] = None
                            _RUN_CACHE.pop(cache_key, None)
                            upload_status.text = "Aucun fichier déposé."
                            upload_status.style(f"color: {COLORS['ink2']}")
                            _update_run_state()
                            remove_btn.visible = False
                            results.clear()
                            results_stockage.clear()
                            ui.notify("Export retiré.", type="info")

                        remove_btn = ui.button(
                            "Enlever", icon="delete_outline", on_click=_remove_file,
                        ).props("flat dense color=grey-7").classes("text-caption")
                        remove_btn.visible = file_ok

                    async def _on_upload(e):
                        """Sauvegarde l'export à un emplacement stable (par tenant) : il reste.

                        NiceGUI 3.x : l'événement expose ``e.file`` (FileUpload), dont
                        la lecture/sauvegarde est asynchrone (``await e.file.save(...)``).
                        """
                        try:
                            _UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
                            await e.file.save(str(export_path))  # écrase l'ancien → remplacement
                            state["export_path"] = str(export_path)
                            state["export_name"] = e.file.name
                            app.storage.user["reconcil_export_name"] = e.file.name
                            upload_status.text = f"✓ {e.file.name} prêt (conservé)."
                            upload_status.style(f"color: {COLORS['success']}")
                            remove_btn.visible = True
                            _update_run_state()
                            _schedule_run()
                        except Exception as exc:  # noqa: BLE001
                            _log.exception("Upload export EB échoué")
                            upload_status.text = f"Erreur lecture fichier : {exc}"
                            upload_status.style(f"color: {COLORS['error']}")

                    with export_box:
                        ui.upload(
                            on_upload=_on_upload,
                            auto_upload=True,
                            max_files=1,
                            label="Déposer / remplacer l'export Easy Beer",
                        ).props('accept=".xlsx" flat bordered').classes("w-full")

                    # Cache des factures Pennylane (partagé, jamais expiré)
                    with ui.row().classes("items-center gap-2"):
                        cache_label = ui.label().classes("text-caption").style(
                            f"color: {COLORS['ink2']}"
                        )

                        def _refresh_cache_stats():
                            n, octets = _CACHE.stats()
                            cache_label.text = (
                                f"Cache factures : {n} facture(s), {octets / 1_048_576:.1f} Mo"
                            )

                        _refresh_cache_stats()

                        def _clear_cache():
                            n = _CACHE.clear()
                            _refresh_cache_stats()
                            ui.notify(f"Cache vidé ({n} factures).", type="info")

                        clear_dlg, _clear_msg, clear_action = confirm_dialog(
                            "Vider le cache des factures ?",
                            "Toutes les factures devront être retéléchargées et "
                            "reparsées au prochain lancement.",
                            "Vider", action_icon="delete_sweep", danger=True,
                        )
                        clear_action.on_click(lambda: (clear_dlg.close(), _clear_cache()))
                        ui.button(
                            "Vider le cache", icon="delete_sweep", on_click=clear_dlg.open,
                        ).props("flat dense color=grey-7").classes("text-caption")

                _update_run_state()

        # ── Onglets Transport / Stockage + zone de progression partagée ────
        with ui.tabs().props("dense align=left").classes("w-full") as cost_tabs:
            tab_transport = ui.tab("Transport", icon="local_shipping")
            tab_stockage = ui.tab("Stockage", icon="warehouse")

        progress_box = ui.column().classes("w-full")

        with ui.tab_panels(cost_tabs, value=tab_transport).classes("w-full"):
            with ui.tab_panel(tab_transport).classes("q-pa-none"):
                results = ui.column().classes("w-full gap-4")
            with ui.tab_panel(tab_stockage).classes("q-pa-none"):
                results_stockage = ui.column().classes("w-full gap-4")

        # ── Logique d'exécution ────────────────────────────────────────────
        # Jeton d'annulation : changer la période/source pendant un calcul rend
        # le calcul en cours obsolète — son callback de progression lève
        # _Cancelled (le thread s'arrête au prochain callback), et son résultat
        # éventuel est jeté au retour.
        run_state: dict = {"token": 0, "prog_timer": None}

        class _Cancelled(Exception):
            pass

        async def _run(manual: bool = False):
            source = source_radio.value
            if source == "export" and (
                not state.get("export_path") or not os.path.exists(state["export_path"])
            ):
                if manual:
                    ui.notify("Déposez d'abord l'export Easy Beer.", type="warning")
                return
            run_state["token"] += 1
            token = run_state["token"]
            if run_state["prog_timer"]:
                run_state["prog_timer"].cancel()
            d1 = (date_min.value or "").strip() or None
            d2 = (date_max.value or "").strip() or None
            state["date_min"], state["date_max"] = d1, d2
            app.storage.user["reconcil_date_min"] = d1
            app.storage.user["reconcil_date_max"] = d2

            results.clear()
            results_stockage.clear()
            progress_box.clear()
            run_btn.disable()

            # Progression réelle : le thread de travail écrit dans ce dict
            # (il ne doit JAMAIS toucher l'UI), un ui.timer le relit côté UI.
            progress = {"done": 0, "total": None, "hits": 0, "phase": None}

            def _progress_cb(done, total, hits):
                if token != run_state["token"]:
                    raise _Cancelled()
                progress.update(done=done, total=total, hits=hits)

            with progress_box:
                with ui.card().classes("w-full").props("flat bordered"):
                    with ui.card_section().classes("q-pa-md column gap-2"):
                        prog_label = ui.label(
                            "Récupération de la liste des factures…"
                        ).classes("text-body2").style(f"color: {COLORS['ink']}")
                        prog_bar = ui.linear_progress(
                            value=0, show_value=False, size="10px",
                        ).props("rounded color=green-8")

            def _tick():
                if progress.get("phase") == "commandes":
                    prog_label.text = "Chargement des commandes via l'API Easy Beer…"
                    prog_bar.props("indeterminate")
                    return
                prog_bar.props(remove="indeterminate")
                total = progress["total"]
                if not total:
                    return
                done, hits = progress["done"], progress["hits"]
                if done == 0:
                    a_charger = max(total - hits, 0)
                    msg = (
                        f"{total} factures sur la période — {hits} en cache, "
                        f"{a_charger} à télécharger"
                    )
                    if a_charger > 20:
                        msg += " (première fois : plusieurs minutes)"
                    prog_label.text = msg
                else:
                    prog_label.text = f"Facture {done}/{total} ({hits} depuis le cache)"
                prog_bar.set_value(done / total)

            prog_timer = ui.timer(0.3, _tick)
            run_state["prog_timer"] = prog_timer

            try:
                from core.reconciliation.io_api import (
                    charger_sources,
                    lire_factures_pennylane,
                )
                from core.reconciliation.io_api_easybeer import lire_commandes_easybeer
                from core.reconciliation.reconciliation_core import reconcilier

                def _work():
                    stocks: list = []  # factures de stockage trouvées dans le même flux
                    if source == "api":
                        # Commandes en direct depuis Easy Beer (1 requête, qq secondes)
                        progress["phase"] = "commandes"
                        commandes = lire_commandes_easybeer()
                        progress["phase"] = None
                        factures = lire_factures_pennylane(
                            date_min=d1, date_max=d2,
                            cache=_CACHE, progress_cb=_progress_cb,
                            stockage_out=stocks,
                        )
                    else:
                        factures, commandes = charger_sources(
                            state["export_path"], date_min=d1, date_max=d2,
                            cache=_CACHE, progress_cb=_progress_cb,
                            stockage_out=stocks,
                        )
                    return reconcilier(factures, commandes), stocks

                res, stocks = await asyncio.to_thread(_work)
            except _Cancelled:
                return  # remplacé par un lancement plus récent
            except Exception as exc:  # noqa: BLE001
                _log.exception("Réconciliation échouée")
                if token == run_state["token"]:
                    progress_box.clear()
                    with progress_box:
                        error_banner(f"Échec de la réconciliation : {exc}")
                return
            finally:
                prog_timer.cancel()
                if token == run_state["token"]:
                    run_btn.enable()

            if token != run_state["token"]:
                return  # résultat obsolète : la période a changé pendant le calcul
            progress_box.clear()
            results.clear()
            results_stockage.clear()
            _refresh_cache_stats()
            _RUN_CACHE[cache_key] = {
                "date_min": d1, "date_max": d2, "result": res, "stockage": stocks,
            }
            _render_results(res)
            _render_stockage(stocks, res)

        async def _run_manual():
            await _run(manual=True)

        run_btn.on_click(_run_manual)

        # ── Relance automatique : debounce 1 s après la dernière modification ──
        debounce: dict = {"timer": None}

        def _schedule_run():
            if debounce["timer"]:
                debounce["timer"].cancel()
            # immediate=False : NiceGUI déclenche sinon le callback tout de suite,
            # même avec once=True — ce qui annulerait l'effet debounce.
            debounce["timer"] = ui.timer(1.0, _run, once=True, immediate=False)

        date_min.on_value_change(lambda e: _schedule_run())
        date_max.on_value_change(lambda e: _schedule_run())

        # ── Rendu des résultats (3 niveaux) ────────────────────────────────
        def _render_results(res):
            k = res.kpis
            eb_headers = _eb_headers(res)
            with results:
                # ─── a) Synthèse (KPIs) ───────────────────────────────────
                section_title("Synthèse", "insights")
                with ui.row().classes("w-full gap-3 wrap reconcil-kpis"):
                    kpi_card("euro", "Coût transport total",
                             _eur(k.get("cout_transport_total_eur")))
                    kpi_card("scale", "Coût moyen €/kg",
                             _eur_kg(k.get("cout_moyen_eur_par_kg")), COLORS["blue"])
                    kpi_card("percent", "Part transport / HT",
                             _pct(k.get("part_transport_dans_ht")), COLORS["blue"])
                    kpi_card("balance", "Écart poids total",
                             _kg(k.get("ecart_poids_total_kg")), COLORS["orange"])
                    kpi_card("local_shipping", "Livraisons appariées",
                             str(k.get("livraisons_appariees", 0)))
                    kpi_card("warning", "Écarts notables",
                             str(k.get("ecarts_notables", 0)), COLORS["orange"])
                    kpi_card("error_outline", "À vérifier (négatif)",
                             str(k.get("lignes_a_verifier_negatif", 0)), COLORS["error"])
                    kpi_card("help_outline", "Sans pièce (non rapprochées)",
                             str(k.get("lignes_sans_piece", 0)), COLORS["ink2"])

                if not res.lignes:
                    ui.label(
                        "Aucune ligne rapprochée sur cette période."
                    ).classes("text-body2 q-mt-md").style(f"color: {COLORS['ink2']}")
                    return

                # ─── b) Tableau décisionnel ───────────────────────────────
                section_title(f"Tableau décisionnel ({len(res.lignes)} lignes)", "table_rows")
                deci_columns = [
                    {"name": "numero", "label": "N° cmd", "field": "numero",
                     "align": "left", "sortable": True},
                    {"name": "client", "label": "Client", "field": "client",
                     "align": "left", "sortable": True},
                    {"name": "poids_eb", "label": "Poids EB", "field": "poids_eb",
                     "align": "right"},
                    {"name": "poids_sof", "label": "Poids SOFRIPA", "field": "poids_sof",
                     "align": "right"},
                    {"name": "ecart_kg", "label": "Écart kg", "field": "ecart_kg",
                     "align": "right"},
                    {"name": "ecart_pct", "label": "Écart %", "field": "ecart_pct",
                     "align": "right"},
                    {"name": "cout", "label": "Coût transport", "field": "cout",
                     "align": "right"},
                    {"name": "montant_ht", "label": "Montant HT", "field": "montant_ht",
                     "align": "right"},
                    {"name": "transport_ht", "label": "Transport/HT", "field": "transport_ht",
                     "align": "right"},
                    {"name": "eur_kg", "label": "€/kg facturé", "field": "eur_kg",
                     "align": "right"},
                    {"name": "statut", "label": "Statut", "field": "statut",
                     "align": "left", "sortable": True},
                ]
                deci_rows = [
                    {
                        "numero": L.numero,
                        "client": L.client or "—",
                        "poids_eb": _kg(L.poids_eb),
                        "poids_sof": _kg(L.poids_sofripa),
                        "ecart_kg": _kg(L.ecart_kg),
                        "ecart_pct": _pct(L.ecart_pct),
                        "cout": _eur(L.cout_transport),
                        "montant_ht": _eur(L.montant_ht),
                        "transport_ht": _pct(L.transport_sur_ht),
                        "eur_kg": _eur_kg(L.eur_par_kg),
                        "statut": L.statut,
                    }
                    for L in res.lignes
                ]
                deci = ui.table(
                    columns=deci_columns, rows=deci_rows, row_key="numero",
                    pagination={"rowsPerPage": 50},
                ).classes("w-full").props(
                    'flat bordered dense :rows-per-page-options="[25,50,100]"'
                )
                deci.add_slot("body-cell-statut", _STATUT_BADGE_JS)

                # ─── Sans pièce : suggestions à vérifier ──────────────────
                if res.sans_piece:
                    section_title(
                        f"Sans pièce — non rapprochées ({len(res.sans_piece)})",
                        "help_outline",
                    )
                    ui.label(
                        "Lignes facturées par SOFRIPA sans N° pièce exploitable : "
                        "impossibles à rattacher automatiquement à une commande. "
                        "La colonne « Suggestion » est une simple supposition "
                        "(même client, date d'expédition proche, poids cohérent) — "
                        "à vérifier à la main, rien n'est garanti. Ces lignes ne "
                        "comptent ni dans les KPIs ni dans les autres tableaux."
                    ).classes("text-caption q-mb-xs").style(f"color: {COLORS['ink2']}")

                    suggestions = getattr(res, "sans_piece_suggestions", None) or [
                        None
                    ] * len(res.sans_piece)
                    sp_columns = [
                        {"name": "date", "label": "Date exp.", "field": "date",
                         "align": "left"},
                        {"name": "ot", "label": "N° OT", "field": "ot", "align": "left"},
                        {"name": "client", "label": "Client (facture)", "field": "client",
                         "align": "left"},
                        {"name": "poids", "label": "Poids facturé", "field": "poids",
                         "align": "right"},
                        {"name": "cout", "label": "Coût transport", "field": "cout",
                         "align": "right"},
                        {"name": "sugg", "label": "Suggestion Easy Beer (à vérifier)",
                         "field": "sugg", "align": "left"},
                    ]
                    sp_rows = [
                        {
                            "date": f.exp_date or "—",
                            "ot": f.ot or "—",
                            "client": f.client or "—",
                            "poids": _kg(f.poids),
                            "cout": _eur(f.montant),
                            "sugg": (
                                f"EB {c.numero} — {c.client} (à vérifier)"
                                if c is not None else "aucune piste"
                            ),
                        }
                        for f, c in zip(res.sans_piece, suggestions)
                    ]
                    ui.table(
                        columns=sp_columns, rows=sp_rows, row_key="ot",
                        pagination={"rowsPerPage": 50},
                    ).classes("w-full").props(
                        'flat bordered dense :rows-per-page-options="[25,50,100]"'
                    )

                # ─── c) Tableau général (toutes les colonnes) ─────────────
                with ui.row().classes("w-full items-center justify-between q-mt-md"):
                    section_title(
                        f"Tableau général — toutes les données "
                        f"({len(eb_headers) + len(_CALC_COLS)} colonnes)", "grid_on",
                    )

                    async def _export():
                        data = await asyncio.to_thread(_build_mega_xlsx, res, eb_headers)
                        dmin = state.get("date_min") or "debut"
                        dmax = state.get("date_max") or "fin"
                        ui.download.content(
                            data, f"reconciliation_{dmin}_{dmax}.xlsx",
                            media_type=(
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet"
                            ),
                        )

                    ui.button("Exporter en Excel", icon="download", on_click=_export).props(
                        "outline color=green-8"
                    )

                mega_labels = eb_headers + [c[0] for c in _CALC_COLS]
                n_eb = len(eb_headers)
                mega_columns = [
                    {
                        "name": f"c{i}", "label": lab, "field": f"c{i}", "align": "left",
                        "headerClasses": "mega-eb-head" if i < n_eb else "mega-calc-head",
                    }
                    for i, lab in enumerate(mega_labels)
                ]
                mega_rows = []
                for L in res.lignes:
                    brut = L.commande.brut if (L.commande and L.commande.brut) else {}
                    row = {}
                    for i, h in enumerate(eb_headers):
                        row[f"c{i}"] = _eb_display(brut.get(h))
                    for j, (_, getter, disp, _xl) in enumerate(_CALC_COLS):
                        row[f"c{n_eb + j}"] = disp(getter(L))
                    mega_rows.append(row)

                mega = ui.table(
                    columns=mega_columns, rows=mega_rows, row_key="c0",
                    pagination={"rowsPerPage": 50},
                ).classes("w-full mega-table").props(
                    'flat bordered dense :rows-per-page-options="[25,50,100]"'
                )
                # Badge coloré du statut (même couleurs que le tableau décisionnel),
                # sur la dernière colonne (Statut poids).
                statut_key = f"c{n_eb + len(_CALC_COLS) - 1}"
                mega.add_slot(f"body-cell-{statut_key}", _STATUT_BADGE_JS)

                # ─── Transferts internes (SYMBIOSE KEFIR) — exclus des tableaux ───
                section_title(
                    f"Transferts internes — exclus ({len(res.internes)})", "swap_horiz",
                )
                ui.label(
                    "Enlèvements de bouteilles Symbiose (destinataire « SYMBIOSE KEFIR… ») : "
                    "exclus de la réconciliation, regroupés ici."
                ).classes("text-caption q-mb-xs").style(f"color: {COLORS['ink2']}")
                if res.internes:
                    int_columns = [
                        {"name": "date", "label": "Date exp.", "field": "date", "align": "left"},
                        {"name": "ot", "label": "N° OT", "field": "ot", "align": "left"},
                        {"name": "dest", "label": "Destination", "field": "dest", "align": "left"},
                        {"name": "cout", "label": "Coût transport (€)", "field": "cout",
                         "align": "right"},
                    ]
                    int_rows = [
                        {
                            "date": f.exp_date or "—",
                            "ot": f.ot or "—",
                            "dest": f.client or "—",
                            "cout": _eur(f.montant),
                        }
                        for f in res.internes
                    ]
                    total_int = sum((f.montant or 0) for f in res.internes)
                    int_rows.append(
                        {"date": "", "ot": "", "dest": "TOTAL", "cout": _eur(total_int)}
                    )
                    ui.table(
                        columns=int_columns, rows=int_rows, row_key="ot",
                        pagination={"rowsPerPage": 50},
                    ).classes("w-full").props(
                        'flat bordered dense :rows-per-page-options="[25,50,100]"'
                    )
                else:
                    ui.label("Aucun transfert interne sur cette période.").classes(
                        "text-body2"
                    ).style(f"color: {COLORS['ink2']}")

                # ─── Synthèse par enseigne (triée alphabétiquement) ───────────
                section_title("Synthèse par enseigne", "store")
                groupes = sorted(res.par_enseigne, key=lambda g: (g.enseigne or "").lower())
                ens_columns = [
                    {"name": "ens", "label": "Enseigne / tournée", "field": "ens",
                     "align": "left"},
                    {"name": "nb", "label": "Nb livraisons", "field": "nb", "align": "right"},
                    {"name": "poids", "label": "Poids SOFRIPA (kg)", "field": "poids",
                     "align": "right"},
                    {"name": "ecart", "label": "Écart poids (kg)", "field": "ecart",
                     "align": "right"},
                    {"name": "cout", "label": "Coût transport (€)", "field": "cout",
                     "align": "right"},
                    {"name": "ht", "label": "Montant HT (€)", "field": "ht", "align": "right"},
                    {"name": "eurkg", "label": "€ / kg", "field": "eurkg", "align": "right"},
                    {"name": "thp", "label": "Transport / HT", "field": "thp", "align": "right"},
                ]
                ens_rows = [
                    {
                        "ens": g.enseigne,
                        "nb": g.nb_livraisons,
                        "poids": _kg(g.poids_sofripa),
                        "ecart": _kg(g.ecart_kg),
                        "cout": _eur(g.cout_transport),
                        "ht": _eur(g.montant_ht),
                        "eurkg": _eur_kg(g.eur_par_kg),
                        "thp": _pct(g.transport_sur_ht),
                    }
                    for g in groupes
                ]
                if groupes:
                    ens_rows.append({
                        "ens": "TOTAL",
                        "nb": sum(g.nb_livraisons for g in groupes),
                        "poids": _kg(sum(g.poids_sofripa for g in groupes)),
                        "ecart": _kg(sum(g.ecart_kg for g in groupes)),
                        "cout": _eur(sum(g.cout_transport for g in groupes)),
                        "ht": _eur(sum(g.montant_ht for g in groupes)),
                        "eurkg": "—",
                        "thp": "—",
                    })
                ui.table(
                    columns=ens_columns, rows=ens_rows, row_key="ens",
                    pagination={"rowsPerPage": 0},
                ).classes("w-full").props("flat bordered dense")

        # ── Rendu de l'onglet Stockage ──────────────────────────────────────
        def _render_stockage(stocks, res):
            with results_stockage:
                if not stocks:
                    ui.label(
                        "Aucune facture de stockage sur la période sélectionnée. "
                        "Les factures de stockage SOFRIPA sont mensuelles et datées "
                        "en fin de mois — élargissez la période si besoin."
                    ).classes("text-body2 q-mt-md").style(f"color: {COLORS['ink2']}")
                    return

                syn = synthese_stockage(stocks, res.kpis if res else None)

                # ── Synthèse stockage (KPIs) ──────────────────────────────
                section_title("Synthèse stockage", "warehouse")
                with ui.row().classes("w-full gap-3 wrap reconcil-kpis"):
                    kpi_card("warehouse", "Coût stockage total (HT)",
                             _eur(syn["total_ht"]))
                    kpi_card("receipt_long", "Total TTC",
                             _eur(syn["total_ttc"]), COLORS["ink2"])
                    kpi_card("calendar_month", "Moyenne mensuelle (HT)",
                             _eur(syn["moyenne_mensuelle_ht"]), COLORS["blue"])
                    if syn.get("cout_logistique_ht") is not None:
                        kpi_card("summarize", "Coût logistique total (HT)",
                                 _eur(syn["cout_logistique_ht"]))
                        kpi_card("pie_chart", "Part stockage / logistique",
                                 _pct(syn["part_stockage"]), COLORS["orange"])
                        kpi_card("scale", "Stockage / kg expédié",
                                 _eur_kg(syn["stockage_par_kg"]), COLORS["blue"])
                        kpi_card("local_shipping", "Stockage / livraison",
                                 _eur(syn["stockage_par_livraison"]), COLORS["blue"])
                        kpi_card("percent", "Stockage / CA HT",
                                 _pct(syn["part_stockage_sur_ca"]), COLORS["ink2"])

                # ── Factures de stockage par mois (+ évolution) ───────────
                section_title(f"Factures de stockage ({len(stocks)})", "receipt")
                stk_columns = [
                    {"name": "periode", "label": "Période", "field": "periode",
                     "align": "left"},
                    {"name": "date", "label": "Date facture", "field": "date",
                     "align": "left"},
                    {"name": "ht", "label": "HT", "field": "ht", "align": "right"},
                    {"name": "tva", "label": "TVA", "field": "tva", "align": "right"},
                    {"name": "ttc", "label": "TTC", "field": "ttc", "align": "right"},
                    {"name": "evol", "label": "Évol. HT vs mois préc.", "field": "evol",
                     "align": "right"},
                ]
                stocks_tries = sorted(stocks, key=lambda s: s.get("date") or "")
                stk_rows, prev = [], None
                for s in stocks_tries:
                    evol = ((s["ht"] - prev) / prev) if prev else None
                    stk_rows.append({
                        "periode": (
                            f"Du {s['periode'].lower()}" if s.get("periode")
                            else (s.get("date") or "—")
                        ),
                        "date": s.get("date") or "—",
                        "ht": _eur(s["ht"]),
                        "tva": _eur(s["tva"]),
                        "ttc": _eur(s["ttc"]),
                        "evol": _pct(evol) if evol is not None else "—",
                    })
                    prev = s["ht"]
                stk_rows.append({
                    "periode": "TOTAL", "date": "",
                    "ht": _eur(syn["total_ht"]),
                    "tva": _eur(round(sum(x["tva"] for x in stocks), 2)),
                    "ttc": _eur(syn["total_ttc"]),
                    "evol": "",
                })
                ui.table(
                    columns=stk_columns, rows=stk_rows, row_key="periode",
                    pagination={"rowsPerPage": 0},
                ).classes("w-full").props("flat bordered dense")

                # ── Répartition indicative par enseigne ───────────────────
                rep = repartition_par_enseigne(
                    syn["total_ht"], res.par_enseigne if res else [],
                )
                if rep:
                    section_title("Répartition par enseigne (indicative)", "store")
                    ui.label(
                        "Le stockage est facturé globalement (une ligne par mois) : "
                        "cette ventilation le répartit au prorata du poids expédié "
                        "par enseigne sur la période. C'est un modèle indicatif, "
                        "pas une donnée facturée."
                    ).classes("text-caption q-mb-xs").style(f"color: {COLORS['ink2']}")
                    rep_columns = [
                        {"name": "ens", "label": "Enseigne / tournée", "field": "ens",
                         "align": "left"},
                        {"name": "poids", "label": "Poids expédié", "field": "poids",
                         "align": "right"},
                        {"name": "part", "label": "% du poids", "field": "part",
                         "align": "right"},
                        {"name": "alloue", "label": "Stockage alloué (HT)",
                         "field": "alloue", "align": "right"},
                    ]
                    rep_rows = [
                        {
                            "ens": r["enseigne"],
                            "poids": _kg(r["poids"]),
                            "part": _pct(r["part"]),
                            "alloue": _eur(r["alloue"]),
                        }
                        for r in rep
                    ]
                    rep_rows.append({
                        "ens": "TOTAL",
                        "poids": _kg(sum(r["poids"] for r in rep)),
                        "part": _pct(1.0),
                        "alloue": _eur(syn["total_ht"]),
                    })
                    ui.table(
                        columns=rep_columns, rows=rep_rows, row_key="ens",
                        pagination={"rowsPerPage": 0},
                    ).classes("w-full").props("flat bordered dense")

        # ── Restauration / premier chargement ──────────────────────────────
        if cached and cached.get("result") is not None:
            # Dernier résultat connu → réaffiché immédiatement, sans relancer.
            _render_results(cached["result"])
            _render_stockage(cached.get("stockage") or [], cached["result"])
        else:
            # Premier passage : lancement automatique sur la période par défaut
            # (mois précédent) — l'utilisateur arrive, les chiffres se chargent.
            ui.timer(0.2, _run, once=True)
