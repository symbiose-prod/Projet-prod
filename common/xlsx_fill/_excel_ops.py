"""
common/xlsx_fill/_excel_ops.py
==============================
Safe Excel cell writing (handles merged cells) and image insertion.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
except ImportError:
    AnchorMarker = TwoCellAnchor = None

logger = logging.getLogger("ferment.xlsx_fill")


# ======= Ecriture suere dans les fusions =====================================

def _safe_set_cell(ws, row: int, col: int, value, number_format: str | None = None):
    """
    Ecrit une valeur *meme si* (row,col) tombe dans une cellule fusionnee.
    """
    hit = None
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            hit = rng
            break

    if hit is None:
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if number_format:
            cell.number_format = number_format
        return

    a_row, a_col = hit.min_row, hit.min_col
    coord = hit.coord
    if row == a_row and col == a_col:
        cell = ws.cell(row=a_row, column=a_col)
        cell.value = value
        if number_format:
            cell.number_format = number_format
        return

    ws.unmerge_cells(coord)
    cell = ws.cell(row=a_row, column=a_col)
    cell.value = value
    if number_format:
        cell.number_format = number_format
    ws.merge_cells(coord)


def _set(ws, addr: str, value, number_format: str | None = None):
    """Ecrit via adresse A1 en gerant les fusions."""
    row, col = coordinate_to_tuple(addr)
    _safe_set_cell(ws, row, col, value, number_format)
    return f"{get_column_letter(col)}{row}"


def _addr(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"


# ======= Logo ================================================================

def _add_logo(ws, path: Path | None, anchor_cell: str, max_w: int, max_h: int):
    """Ajoute un logo ancre sans deformer l'image (no-op si chemin invalide)."""
    try:
        if not path or not path.exists():
            logger.debug("Logo introuvable pour ancre %s -> %s", anchor_cell, path)
            return
        from PIL import Image as PILImage

        with PILImage.open(path) as im:
            ow, oh = im.size

        scale = min(max_w / ow, max_h / oh, 1.0)
        img = XLImage(str(path))
        img.width = max(1, int(round(ow * scale)))
        img.height = max(1, int(round(oh * scale)))
        ws.add_image(img, anchor_cell)
        logger.debug("Logo ajoute: %s -> %s (%dx%dpx)", path.name, anchor_cell, img.width, img.height)
    except Exception as e:
        logger.error("Erreur ajout logo %s @ %s: %s", path, anchor_cell, e)


# ======= Insertion d'image ancree ============================================

def _add_image_in_range(ws, img_path: Path, tl_addr: str, br_addr: str):
    """Insere une image et l'ancre sur la plage [tl_addr:br_addr]."""
    try:
        if not img_path or not img_path.exists():
            logger.debug("Image introuvable: %s", img_path)
            return

        img = XLImage(str(img_path))
        logger.debug("Image OK: %s", img_path.name)

        # 1) Tentative TwoCellAnchor (precise)
        if AnchorMarker and TwoCellAnchor:
            try:
                tl_row, tl_col = coordinate_to_tuple(tl_addr)
                br_row, br_col = coordinate_to_tuple(br_addr)
                frm = AnchorMarker(col=tl_col - 1, colOff=0, row=tl_row - 1, rowOff=0)
                to = AnchorMarker(col=br_col - 1, colOff=0, row=br_row - 1, rowOff=0)
                img.anchor = TwoCellAnchor(_from=frm, _to=to, editAs="oneCell")
                ws.add_image(img)
                logger.debug("Image ancree via TwoCellAnchor.")
                return
            except Exception as e:
                logger.debug("TwoCellAnchor indisponible/echec: %s", e)

        # 2) Fallback: ancre en coin superieur gauche
        ws.add_image(img, tl_addr)
        logger.debug("Image ajoutee en %s (fallback). Redimension approx...", tl_addr)

        def _col_pixels(col_idx_1b: int) -> int:
            col_letter = get_column_letter(col_idx_1b)
            w = ws.column_dimensions[col_letter].width
            if w is None:
                w = 8.43
            return int(round(w * 7.0))

        def _row_pixels(row_idx_1b: int) -> int:
            h = ws.row_dimensions[row_idx_1b].height
            if h is None:
                h = 15
            return int(round(h * (96.0 / 72.0)))

        tl_r, tl_c = coordinate_to_tuple(tl_addr)
        br_r, br_c = coordinate_to_tuple(br_addr)
        width_px = sum(_col_pixels(c) for c in range(tl_c, br_c + 1))
        height_px = sum(_row_pixels(r) for r in range(tl_r, br_r + 1))
        if width_px > 0 and height_px > 0:
            img.width, img.height = width_px, height_px
            logger.debug("Redimension: %dx%dpx", width_px, height_px)
    except Exception as e:
        logger.error("Erreur insertion image: %s", e)
