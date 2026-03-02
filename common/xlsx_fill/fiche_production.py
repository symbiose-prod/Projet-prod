"""
common/xlsx_fill/fiche_production.py
====================================
Fill production sheet (Fiche_production.xlsx).
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date
from typing import Optional, Dict

import pandas as pd

_log = logging.getLogger("ferment.xlsx_fill")

from ._helpers import _project_root, _to_excel_label, FILTRE_RATIO_KEFIR
from ._excel_ops import _safe_set_cell, _set, _add_logo
from ._stock_parse import _parse_format_from_stock
from ._tank_ruler import interpolate_ruler_height


def fill_fiche_xlsx(
    template_path: str,
    semaine_du: date,
    ddm: date,
    gout1: str,
    gout2: Optional[str] = None,
    df_calc=None,
    sheet_name: str | None = None,
    df_min=None,
    *,
    V_start: float = 0.0,
    tank_capacity: int = 7200,
    transfer_loss: float = 400.0,
    aromatisation_volume: float = 0.0,
    is_infusion: bool = False,
    dilution_ingredients: Dict[str, float] | None = None,
) -> bytes:
    """
    Remplit la fiche de production unique (Fiche_production.xlsx).
    """
    import openpyxl

    # --- ouverture & selection de la feuille ---
    wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=False)
    targets = [sheet_name] if sheet_name else ["Fiche de production"]
    ws = None
    for nm in targets:
        if nm and nm in wb.sheetnames:
            ws = wb[nm]
            break
    if ws is None:
        ws = wb.active

    # --- Mise en page ---
    try:
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = 100
        ws.page_setup.horizontalCentered = True
    except Exception:
        _log.debug("Erreur mise en forme cellule", exc_info=True)

    # --- Logos ---
    root = _project_root()

    def _first_existing(paths):
        for p in paths:
            if p.exists():
                return p
        return None

    symbiose_path = _first_existing([
        root / "assets" / "logo_symbiose.png",
        root / "assets" / "signature" / "logo_symbiose.png",
        root / "assets" / "Logo_Symbiose.png",
    ])
    niko_path = _first_existing([
        root / "assets" / "NIKO_Logo.png",
        root / "assets" / "signature" / "NIKO_Logo.png",
        root / "assets" / "niko_logo.png",
    ])

    _add_logo(ws, symbiose_path, anchor_cell="A1", max_w=80, max_h=55)
    _add_logo(ws, niko_path, anchor_cell="B1", max_w=65, max_h=45)

    # --- Titre "Cuve de xxxxL" ---
    if tank_capacity > 0:
        from openpyxl.styles import Font as _Font, Alignment as _Align
        _set(ws, "C1", f"Cuve de {tank_capacity}L")
        try:
            ws["C1"].font = _Font(name="Aptos Narrow", size=20, bold=True)
            ws["C1"].alignment = _Align(horizontal="center", vertical="center")
        except Exception:
            _log.debug("Erreur mise en forme DDM", exc_info=True)

    # --- A21 : date de debut de production ---
    _set(ws, "A21", semaine_du, number_format="DD/MM/YYYY")

    # --- B8 : gout ---
    _set(ws, "B8", _to_excel_label(gout1) or "")

    # --- DDM : B10:C10 ---
    try:
        from openpyxl.styles import Alignment
        for rng in list(ws.merged_cells.ranges):
            if not (rng.max_row < 10 or rng.min_row > 10 or rng.max_col < 2 or rng.min_col > 3):
                ws.unmerge_cells(rng.coord)
        ws.merge_cells("B10:C10")
        _safe_set_cell(ws, 10, 2, ddm, number_format="DD/MM/YYYY")
        ws["B10"].alignment = Alignment(vertical="center", horizontal="left")
    except Exception:
        _log.debug("Erreur merge_cells, tentative alternative", exc_info=True)
        try:
            ws.merge_cells("B10:C10")
        except Exception:
            _log.debug("Erreur merge_cells alternative", exc_info=True)
        _safe_set_cell(ws, 10, 2, ddm, number_format="DD/MM/YYYY")

    # --- Rows 15-16 : bouteilles et cartons par format ---
    SLOT_COL = {
        "sym_33_x12":  2,
        "sym_33_x6":   3,
        "niko_33_x12": 4,
        "inter_33_x6": 5,
        "sym_75_x6":   6,
        "sym_75_x4":   7,
        "niko_75_x6":  8,
        "autre_75":    9,
    }
    cartons_by_slot = {k: 0 for k in SLOT_COL}
    bouteilles_by_slot = {k: 0 for k in SLOT_COL}

    if isinstance(df_min, pd.DataFrame) and not df_min.empty:
        dff = df_min.copy()
        if "GoutCanon" in dff.columns:
            dff = dff[dff["GoutCanon"].astype(str).str.strip() == str(gout1 or "").strip()]

        col_cart = next((c for c in dff.columns if "Cartons" in str(c) and "produire" in str(c)), None)
        col_btl = next((c for c in dff.columns if "Bouteilles" in str(c) and "produire" in str(c)), None)

        if col_cart and not dff.empty:
            for _, r0 in dff.iterrows():
                ct = int(pd.to_numeric(r0.get(col_cart), errors="coerce") or 0)
                bt = int(pd.to_numeric(r0.get(col_btl), errors="coerce") or 0) if col_btl else 0
                if ct <= 0:
                    continue

                prod = str(r0.get("Produit", "")).upper()
                stock = str(r0.get("Stock", ""))
                nb, volL = _parse_format_from_stock(stock)

                if nb is None or volL is None:
                    m_nb = re.search(r"x\s*(\d+)", prod)
                    m_vol = re.search(r"(\d+(?:[.,]\d+)?)\s*cL", prod, flags=re.I)
                    nb = int(m_nb.group(1)) if m_nb else nb
                    volL = (float(m_vol.group(1).replace(",", ".")) / 100.0) if m_vol else volL

                if nb is None or volL is None:
                    continue

                has_niko = "NIKO" in prod
                is_inter = "PROBIOTIC" in prod or "WATER KEFIR" in prod

                if abs(volL - 0.33) < 0.01:
                    if nb == 12:
                        slot = "niko_33_x12" if has_niko else "sym_33_x12"
                    elif nb == 6:
                        slot = "inter_33_x6" if is_inter else "sym_33_x6"
                    else:
                        slot = "sym_33_x12"
                elif abs(volL - 0.75) < 0.01:
                    if nb == 6:
                        slot = "niko_75_x6" if has_niko else "sym_75_x6"
                    elif nb == 4:
                        slot = "sym_75_x4"
                    else:
                        slot = "autre_75"
                else:
                    continue

                cartons_by_slot[slot] += ct
                bouteilles_by_slot[slot] += bt

    for slot_name, col_idx in SLOT_COL.items():
        bt = bouteilles_by_slot[slot_name]
        ct = cartons_by_slot[slot_name]
        if bt > 0:
            _safe_set_cell(ws, 15, col_idx, bt)
        if ct > 0:
            _safe_set_cell(ws, 16, col_idx, ct)

    # --- C30-C33 : ingredients dilution ---
    if dilution_ingredients:
        DILUTION_CELLS = {
            "sucre": "C30",
            "figue": "C31",
            "citron": "C32",
            "grain": "C33",
        }
        _used_cells: set[str] = set()
        for libelle, qty in dilution_ingredients.items():
            lib_lower = libelle.lower()
            matched = False
            for keyword, cell_addr in DILUTION_CELLS.items():
                if keyword in lib_lower and cell_addr not in _used_cells:
                    _set(ws, cell_addr, round(qty, 2))
                    _used_cells.add(cell_addr)
                    matched = True
                    break
            if not matched:
                for _fb_addr in ["C30", "C31", "C32", "C33"]:
                    if _fb_addr not in _used_cells:
                        _set(ws, _fb_addr, round(qty, 2))
                        _used_cells.add(_fb_addr)
                        break

    # --- C35-C36 : volume de remplissage + niveau de liquide ---
    if V_start > 0:
        _set(ws, "C35", round(V_start))
        if tank_capacity > 0:
            h_start = interpolate_ruler_height(V_start, tank_capacity)
            _set(ws, "C36", round(h_start, 1))

    # --- Phase 2 : Filtration (B42-B44) ---
    if V_start > 0 and transfer_loss >= 0:
        V_transferred = V_start - transfer_loss
        filtre_ratio = 0.0 if is_infusion else FILTRE_RATIO_KEFIR

        volume_filtre = V_transferred * filtre_ratio
        volume_non_filtre = V_transferred - volume_filtre
        volume_final = volume_filtre + aromatisation_volume

        _set(ws, "B42", round(volume_filtre))
        _set(ws, "B43", round(volume_final))
        if tank_capacity > 0:
            h_final = interpolate_ruler_height(volume_final, tank_capacity)
            _set(ws, "B44", round(h_final, 1))

        # --- Phase 2 : Remplissage (B48-B49) ---
        volume_total = volume_final + volume_non_filtre
        _set(ws, "B48", round(volume_total))
        if tank_capacity > 0:
            h_total = interpolate_ruler_height(volume_total, tank_capacity)
            _set(ws, "B49", round(h_total, 1))

    # Sauvegarde en memoire
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
