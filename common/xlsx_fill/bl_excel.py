"""
common/xlsx_fill/bl_excel.py
============================
Fill BL enlevements Excel template.
"""
from __future__ import annotations

import io
import os
import re
import unicodedata
from datetime import date, datetime
from typing import List

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


def fill_bl_enlevements_xlsx(
    template_path: str,
    date_creation: date,
    date_ramasse: date,
    destinataire_title: str,
    destinataire_lines: List[str],
    df_lines: pd.DataFrame,
) -> bytes:
    """
    Remplit le modele XLSX 'LOG_EN_001_01 BL enlevements Sofripa-2.xlsx'
    de facon ANCREE sur la rangee d'en-tetes reelle (sequence contigue).
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Mod\u00e8le Excel introuvable: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ---------- utilitaires locaux ----------
    def _safe_write(ws, row: int, col: int, value):
        r, c = row, col
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                r, c = rng.min_row, rng.min_col
                break
        ws.cell(row=r, column=c).value = value

    def _norm(s: str) -> str:
        s = str(s or "").strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.replace("\u2019", "'")
        for ch in ("(", ")", ":", ";", ","):
            s = s.replace(ch, " ")
        return " ".join(s.split())

    def _find_header_run(ws):
        SEQ = [
            ["reference"],
            ["produit", "produit (gout + format)", "produit gout format"],
            ["ddm", "date de durabilite", "date de durabilite"],
            ["quantite cartons", "quantite cartons", "n cartons", "no cartons", "nb cartons"],
            ["quantite palettes", "quantite palettes", "n palettes", "no palettes", "nb palettes"],
            ["poids palettes (kg)", "poids palettes", "poids (kg)"],
        ]
        maxr = min(ws.max_row or 1, 100)
        maxc = min(ws.max_column or 1, 30)
        best = None
        for r in range(1, maxr + 1):
            found_ref = False
            for c0 in range(1, maxc - 4):
                hv = _norm(ws.cell(row=r, column=c0).value)
                if hv in SEQ[0]:
                    found_ref = True
                    break
            if not found_ref:
                continue
            for c0 in range(1, maxc - 4):
                ok = True
                cols = []
                for k in range(6):
                    hv = _norm(ws.cell(row=r, column=c0 + k).value)
                    if hv not in SEQ[k]:
                        ok = False
                        break
                    cols.append(c0 + k)
                if ok:
                    best = (r, cols)
                    break
        return best

    # ---------- 1) Dates ----------
    def _find_cell_by_regex(ws, pattern: str):
        rx = re.compile(pattern, flags=re.I)
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and rx.search(v):
                    return cell.row, cell.column
        return None, None

    r, c = _find_cell_by_regex(ws, r"date\s+de\s+cr[e\u00e9]ation")
    if r and c:
        _safe_write(ws, r, c + 1, date_creation.strftime("%d/%m/%Y"))

    r, c = _find_cell_by_regex(ws, r"date\s+de\s+rammasse|date\s+de\s+ramasse")
    if r and c:
        _safe_write(ws, r, c + 1, date_ramasse.strftime("%d/%m/%Y"))

    # ---------- 2) Destinataire ----------
    r, c = _find_cell_by_regex(ws, r"destinataire")
    if r and c:
        target_rng = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col > c:
                if target_rng is None or rng.min_col < target_rng.min_col:
                    target_rng = rng

        if target_rng:
            rr, cc = target_rng.min_row, target_rng.min_col
            rr_end, cc_end = target_rng.max_row, target_rng.max_col
        else:
            rr, cc = r, c + 1
            rr_end, cc_end = min(r + 2, ws.max_row), min(c + 6, ws.max_column)
            try:
                ws.merge_cells(start_row=rr, start_column=cc, end_row=rr_end, end_column=cc_end)
            except Exception:
                pass

        text = "\n".join([destinataire_title] + [x for x in (destinataire_lines or []) if str(x).strip()])
        _safe_write(ws, rr, cc, text)
        a = ws.cell(row=rr, column=cc).alignment or Alignment()
        ws.cell(row=rr, column=cc).alignment = Alignment(
            wrap_text=True, vertical="top", horizontal=a.horizontal or "left"
        )

        n_lines = max(1, text.count("\n") + 1)
        span = max(1, rr_end - rr + 1)
        per_row = 14 * n_lines / span
        for rset in range(rr, rr_end + 1):
            cur = ws.row_dimensions[rset].height or 0
            ws.row_dimensions[rset].height = max(cur, per_row)

        zr, zc = _find_cell_by_regex(ws, r"zac\s+du\s+haut\s+de\s+wissous")
        if zr and zc and (zr < rr or zr > rr_end or zc < cc or zc > cc_end):
            _safe_write(ws, zr, zc, "")

    # ---------- 3) Localisation en-tetes ----------
    header = _find_header_run(ws)
    if not header:
        raise KeyError("Impossible de localiser la rang\u00e9e d'en-t\u00eates (s\u00e9quence compl\u00e8te non trouv\u00e9e).")
    hdr_row, (c_ref, c_prod, c_ddm, c_qc, c_qp, c_poids) = header

    # ---------- 4) DataFrame normalisation ----------
    df = df_lines.copy()
    if "Produit" not in df.columns and "Produit (go\u00fbt + format)" in df.columns:
        df = df.rename(columns={"Produit (go\u00fbt + format)": "Produit"})

    def _to_ddm_val(x):
        if isinstance(x, date):
            return x.strftime("%d/%m/%Y")
        s = str(x or "").strip()
        if not s:
            return ""
        try:
            if "-" in s and len(s.split("-")[0]) == 4:
                return datetime.strptime(s, "%Y-%m-%d").strftime("%d/%m/%Y")
            return datetime.strptime(s, "%d/%m/%Y").strftime("%d/%m/%Y")
        except Exception:
            return s

    def _as_int(v) -> int:
        try:
            f = float(v)
            return int(round(f))
        except Exception:
            return 0

    # ---------- 5) Ecriture des lignes ----------
    row = hdr_row + 1
    for _, r in df.iterrows():
        _safe_write(ws, row, c_ref, str(r.get("R\u00e9f\u00e9rence", "")))
        _safe_write(ws, row, c_prod, str(r.get("Produit", "")))
        _safe_write(ws, row, c_ddm, _to_ddm_val(r.get("DDM", "")))
        _safe_write(ws, row, c_qc, _as_int(r.get("Quantit\u00e9 cartons", 0)))
        _safe_write(ws, row, c_qp, _as_int(r.get("Quantit\u00e9 palettes", 0)))
        _safe_write(ws, row, c_poids, _as_int(r.get("Poids palettes (kg)", 0)))
        row += 1

    # ---------- 6) Sauvegarde ----------
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
