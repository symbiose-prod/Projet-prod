"""
common/xlsx_export.py
=====================
Export XLSX « 1 ligne par palette (SSCC) » pour l'app mobile.

Deux usages, même format de sortie (chaque ligne = une palette identifiable
par son SSCC, avec la ramasse d'appartenance) :
  - export d'une sélection de ramasses (palettes de ces ramasses),
  - export du journal SSCC (selon les filtres affichés).

On construit le classeur openpyxl « from scratch » (pas de template) : c'est
un export de données tabulaires, pas un document à mise en page fixe comme le
BL. Retour : ``bytes`` prêts à être streamés en réponse HTTP.
"""
from __future__ import annotations

import datetime as _dt
import io
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from common.services.sscc_service import SsccLogEntry

# (clé d'attribut SsccLogEntry, libellé colonne). L'ordre = l'ordre des
# colonnes dans le fichier. Les colonnes « ramasse » d'abord pour qu'on
# identifie tout de suite à quelle ramasse appartient chaque palette.
_COLUMNS: list[tuple[str, str]] = [
    ("ramasse_numero", "Ramasse n°"),
    ("ramasse_date", "Date ramasse"),
    ("ramasse_destinataire", "Destinataire"),
    ("sscc", "SSCC"),
    ("lot", "Lot"),
    ("designation", "Produit"),
    ("marque", "Marque"),
    ("gout", "Goût"),
    ("case_count", "Cartons"),
    ("ddm", "DDM"),
    ("gtin_palette", "GTIN palette"),
    ("generated_at", "Générée le"),
    ("user_email", "Générée par"),
    ("loaded_at", "Chargée le"),
    ("voided", "Annulée"),
    ("voided_reason", "Motif annulation"),
    ("archived", "Archivée"),
    ("label_archived_reason", "Motif archivage"),
]


def _fmt_date(value: Any) -> str:
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    return str(value or "")


def _fmt_datetime(value: Any) -> str:
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    return str(value or "")


def _cell_value(entry: SsccLogEntry, key: str) -> Any:
    """Projette un champ d'une ligne vers la valeur écrite dans la cellule."""
    if key == "ramasse_numero":
        return entry.ramasse_numero if entry.ramasse_numero is not None else ""
    if key == "ramasse_date":
        return _fmt_date(entry.ramasse_date)
    if key == "ddm":
        return _fmt_date(entry.ddm)
    if key == "generated_at":
        return _fmt_datetime(entry.generated_at)
    if key == "loaded_at":
        return _fmt_datetime(entry.loaded_at)
    if key == "case_count":
        return int(entry.case_count or 0)
    if key == "voided":
        return "Oui" if entry.voided_at else "Non"
    if key == "archived":
        return "Oui" if entry.label_archived_at else "Non"
    # Champs texte simples (sscc, lot, designation, marque, gout,
    # gtin_palette, user_email, voided_reason, label_archived_reason).
    return getattr(entry, key, "") or ""


def build_palettes_xlsx(
    entries: list[SsccLogEntry],
    *,
    sheet_title: str = "Palettes",
) -> bytes:
    """Construit un classeur XLSX, 1 ligne par palette (SSCC).

    Args:
        entries: lignes à exporter (``SsccLogEntry`` du service SSCC).
        sheet_title: nom de l'onglet (tronqué à 31 car. — limite Excel).

    Returns:
        Les octets du fichier .xlsx.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Palettes")[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5233")  # vert sombre Symbiose
    header_align = Alignment(vertical="center")

    # En-têtes
    for col_idx, (_key, label) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Lignes de données
    for row_idx, entry in enumerate(entries, start=2):
        for col_idx, (key, _label) in enumerate(_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(entry, key))

    # Largeurs de colonnes : approximées sur le contenu (cap à 48).
    for col_idx, (key, label) in enumerate(_COLUMNS, start=1):
        max_len = len(label)
        for entry in entries:
            max_len = max(max_len, len(str(_cell_value(entry, key))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)

    # Fige la ligne d'en-tête + filtre automatique sur toute la plage.
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(entries) + 1)}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
