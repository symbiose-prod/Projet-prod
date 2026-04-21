"""
core/optimizer/excel_io.py
==========================
Read Excel input from uploaded bytes (Streamlit/NiceGUI upload).
"""
from __future__ import annotations

import io
import logging

import pandas as pd

_log = logging.getLogger("ferment.optimizer.excel_io")

from .parsing import detect_header_row, parse_days_from_b2, rows_to_keep_by_fill

DEFAULT_WINDOW_DAYS = 60


def read_input_excel_and_period_from_bytes(file_bytes: bytes):
    """Meme logique que _from_path mais pour des bytes (uploader)."""
    import openpyxl  # noqa: F401

    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    header_idx = detect_header_row(raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    keep_mask = rows_to_keep_by_fill(file_bytes, header_idx)
    if len(keep_mask) < len(df):
        keep_mask = keep_mask + [True] * (len(df) - len(keep_mask))
    df = df.iloc[[i for i, k in enumerate(keep_mask) if k]].reset_index(drop=True)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        wd = parse_days_from_b2(b2_val)
    except (KeyError, AttributeError, TypeError, ValueError, IndexError):
        _log.debug("Erreur parsing periode depuis B2", exc_info=True)
        wd = None
    return df, (wd if wd and wd > 0 else DEFAULT_WINDOW_DAYS)


def parse_stock_produits_excel(file_bytes: bytes) -> pd.DataFrame:
    """Parse l'export EB ``/stock/produits/export`` en DataFrame compact.

    Colonnes sortie : ``Produit``, ``Stock``, ``Quantité disponible``,
    ``Volume disponible (hl)``.
    Ne garde que les lignes avec un ``Conditionnement`` non vide (une ligne par
    couple produit/format — les lignes agrégées ou par lot DDM sont ignorées).
    Dédoublonne sur (Produit, Stock) en sommant les quantités.
    """
    df = pd.read_excel(io.BytesIO(file_bytes))
    needed = {"Produit", "Conditionnement", "Qté virtuelle", "Vol. virtuel"}
    if not needed.issubset(df.columns):
        _log.warning(
            "stock-produits: colonnes manquantes, attendu %s, reçu %s",
            needed, list(df.columns),
        )
        return pd.DataFrame(columns=["Produit", "Stock", "Quantité disponible", "Volume disponible (hl)"])

    df = df[df["Conditionnement"].notna() & (df["Conditionnement"].astype(str).str.strip() != "")]
    df = df[df["Produit"].notna()]
    df = df[["Produit", "Conditionnement", "Qté virtuelle", "Vol. virtuel"]].rename(
        columns={
            "Conditionnement": "Stock",
            "Qté virtuelle": "Quantité disponible",
            "Vol. virtuel": "Volume disponible (hl)",
        },
    )
    df["Quantité disponible"] = pd.to_numeric(df["Quantité disponible"], errors="coerce").fillna(0.0)
    df["Volume disponible (hl)"] = pd.to_numeric(df["Volume disponible (hl)"], errors="coerce").fillna(0.0)
    # Agrège les lots éventuels par couple produit/format
    df = (
        df.groupby(["Produit", "Stock"], as_index=False, sort=False)
          .agg({"Quantité disponible": "sum", "Volume disponible (hl)": "sum"})
    )
    return df.reset_index(drop=True)


def enrich_df_with_missing_formats(
    df_autonomie: pd.DataFrame, df_stock_produits: pd.DataFrame,
) -> pd.DataFrame:
    """Injecte les couples (Produit, Stock) présents dans le stock mais absents
    de l'export autonomie-stocks (formats à 0 vente sur la période).

    Pour chaque paire manquante, crée une ligne avec ventes=0 et le stock
    virtuel depuis ``df_stock_produits``. L'existant n'est jamais écrasé.
    """
    if df_autonomie is None or df_autonomie.empty:
        return df_autonomie
    if df_stock_produits is None or df_stock_produits.empty:
        return df_autonomie
    if "Produit" not in df_autonomie.columns or "Stock" not in df_autonomie.columns:
        return df_autonomie

    existing = set(
        zip(
            df_autonomie["Produit"].astype(str).str.strip(),
            df_autonomie["Stock"].astype(str).str.strip(),
        ),
    )
    autonomie_products = set(df_autonomie["Produit"].astype(str).str.strip())

    new_rows = []
    for _, r in df_stock_produits.iterrows():
        prod = str(r["Produit"]).strip()
        stock = str(r["Stock"]).strip()
        # Matching produit : le nom dans stock-produits a souvent un suffixe
        # "- 0.0°" que autonomie-stocks omet. On tolère un préfixe commun.
        auto_prod = prod
        if prod not in autonomie_products:
            matched = next(
                (ap for ap in autonomie_products if ap and (prod.startswith(ap) or ap.startswith(prod))),
                None,
            )
            if matched is None:
                continue  # produit inconnu en autonomie — on n'invente pas
            auto_prod = matched
        if (auto_prod, stock) in existing:
            continue
        qte = float(r.get("Quantité disponible", 0) or 0)
        vol = float(r.get("Volume disponible (hl)", 0) or 0)
        # Pas de filtre sur qte/vol : dès qu'EB a enregistré un couple
        # (produit, format), on l'injecte — même à stock 0 ou négatif.
        # Les formats fraîchement créés sans réappro apparaissent ainsi
        # dans le plan et peuvent être forcés manuellement.
        new_rows.append({
            "Produit": auto_prod, "Stock": stock,
            "Quantité vendue": 0, "Volume vendu (hl)": 0.0,
            "Quantité disponible": max(qte, 0.0),
            "Volume disponible (hl)": max(vol, 0.0),
        })

    if not new_rows:
        return df_autonomie

    df_add = pd.DataFrame(new_rows)
    # Aligner les colonnes sur df_autonomie (ajoute les colonnes manquantes en NaN)
    for c in df_autonomie.columns:
        if c not in df_add.columns:
            df_add[c] = pd.NA
    df_add = df_add[df_autonomie.columns]
    _log.info("stock-produits: %d formats injectés depuis l'export stock", len(df_add))
    return pd.concat([df_autonomie, df_add], ignore_index=True)
