"""
core/optimizer/parsing.py
=========================
Stock format parsing, header detection, period parsing.
"""
from __future__ import annotations

import io
import re
from typing import Optional, List

import numpy as np
import pandas as pd


# ======= constantes formats =================================================
ALLOWED_FORMATS = {(12, 0.33), (6, 0.75), (4, 0.75)}
VOL_TOL = 0.02


# ======= parse_stock ========================================================
def parse_stock(text: str):
    if pd.isna(text):
        return np.nan, np.nan
    s = str(text)
    nb = None
    for pat in [
        r"(?:Carton|Caisse|Colis)\s+de\s*(\d+)",
        r"(\d+)\s*[x\u00d7]\s*Bouteilles?",
        r"(\d+)\s*[x\u00d7]",
        r"(\d+)\s+Bouteilles?",
    ]:
        m = re.search(pat, s, flags=re.IGNORECASE)
        if m:
            try:
                nb = int(m.group(1))
                break
            except Exception:
                pass
    vol_l = None
    m_l = re.findall(r"(\d+(?:[.,]\d+)?)\s*[lL]", s)
    if m_l:
        vol_l = float(m_l[-1].replace(",", "."))
    else:
        m_cl = re.findall(r"(\d+(?:[.,]\d+)?)\s*c[lL]", s)
        if m_cl:
            vol_l = float(m_cl[-1].replace(",", ".")) / 100.0
    if nb is None or vol_l is None:
        m_combo = re.search(
            r"(\d+)\s*[x\u00d7]\s*(\d+(?:[.,]\d+)?)+\s*([lc]l?)",
            s, flags=re.IGNORECASE,
        )
        if m_combo:
            try:
                nb2 = int(m_combo.group(1))
                val = float(m_combo.group(2).replace(",", "."))
                unit = m_combo.group(3).lower()
                vol2 = val if unit.startswith("l") else val / 100.0
                if nb is None:
                    nb = nb2
                if vol_l is None:
                    vol_l = vol2
            except Exception:
                pass
    if (nb is None or np.isnan(nb)) and vol_l is not None and abs(vol_l - 0.75) <= VOL_TOL:
        if re.search(r"(?:\b4\s*[x\u00d7]\b|Carton\s+de\s*4\b|4\s+Bouteilles?)", s, flags=re.IGNORECASE):
            nb = 4
    return (float(nb) if nb is not None else np.nan, float(vol_l) if vol_l is not None else np.nan)


def safe_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def is_allowed_format(nb_bottles, vol_l, stock_txt: str) -> bool:
    if pd.isna(nb_bottles) or pd.isna(vol_l):
        if re.search(
            r"(?:\b4\s*[x\u00d7]\s*75\s*c?l\b|\b4\s+Bouteilles?\b.*75\s*c?l)",
            stock_txt, flags=re.IGNORECASE,
        ):
            nb_bottles = 4
            vol_l = 0.75
        else:
            return False
    nb_bottles = int(nb_bottles)
    vol_l = float(vol_l)
    for nb_ok, vol_ok in ALLOWED_FORMATS:
        if nb_bottles == nb_ok and abs(vol_l - vol_ok) <= VOL_TOL:
            return True
    return False


# ======= detection en-tete & periode B2 =====================================
def detect_header_row(df_raw: pd.DataFrame) -> int:
    must = {
        "Produit", "Stock", "Quantité vendue",
        "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)",
    }
    for i in range(min(10, len(df_raw))):
        if must.issubset(set(str(x).strip() for x in df_raw.iloc[i].tolist())):
            return i
    return 0


def rows_to_keep_by_fill(excel_bytes: bytes, header_idx: int) -> List[bool]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    start_row = header_idx + 2
    keep: List[bool] = []
    for r in range(start_row, ws.max_row + 1):
        is_black = False
        for cell in ws[r]:
            fill = cell.fill
            if fill and fill.fill_type:
                rgb = (
                    getattr(getattr(fill, "fgColor", None), "rgb", None)
                    or getattr(getattr(fill, "start_color", None), "rgb", None)
                )
                if rgb and rgb[-6:].upper() == "000000":
                    is_black = True
                    break
        keep.append(not is_black)
    return keep


def parse_days_from_b2(value) -> Optional[int]:
    try:
        if isinstance(value, (int, float)) and not pd.isna(value):
            v = int(round(float(value)))
            return v if v > 0 else None
        if value is None:
            return None
        s = str(value).strip()
        m = re.search(r"(\d+)\s*(?:j|jour|jours)\b", s, flags=re.IGNORECASE)
        if m:
            return int(m.group(1)) or None
        date_pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
        m2 = re.search(date_pat, s)
        if m2:
            d1 = pd.to_datetime(m2.group(1), dayfirst=True, errors="coerce")
            d2 = pd.to_datetime(m2.group(2), dayfirst=True, errors="coerce")
            if pd.notna(d1) and pd.notna(d2):
                days = int((d2 - d1).days)
                return days if days > 0 else None
        m3 = re.search(r"\b(\d{1,4})\b", s)
        if m3:
            v = int(m3.group(1))
            return v if v > 0 else None
    except Exception:
        return None
    return None
