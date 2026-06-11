"""
io_files.py — ENTRÉE "fichiers" (la source actuelle).

Lit la facture transporteur (PDF SOFRIPA) et l'export Easy Beer (Excel),
et les convertit dans le format normalisé attendu par le cœur
(LigneFacture, Commande).

>>> C'est CE module qui sera remplacé par un io_api.py en Brique B
    (Easy Beer + Pennylane), sans toucher au cœur. <<<

Le parsing PDF reprend la logique éprouvée du script auto_reconciliation.py.
"""
import re
from collections import defaultdict

import pdfplumber
from openpyxl import load_workbook

from .reconciliation_core import Commande, LigneFacture

# Bandes de colonnes (positions x) de la facture SOFRIPA — calées sur ce format.
COLS = {'date': (0, 55), 'desig': (55, 300), 'poids': (300, 341), 'colis': (341, 373),
        'palette': (373, 410), 'qte': (410, 456), 'unite': (456, 490),
        'pu': (490, 521), 'montant': (521, 600)}

# Colonnes Easy Beer utiles (on retrouve la bonne colonne par son intitulé).
EB_KEYS = {"num": "N° commande", "client": "Client", "poids": "Poids total",
           "ht": "Total HT", "tournee": "Tournée"}


def _band(x):
    for k, (a, b) in COLS.items():
        if a <= x < b:
            return k


def _num(s):
    s = s.replace(' ', '').replace('\u00a0', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _merge(tokens, col):
    parts = [x for xx, x in sorted(tokens) if _band(xx) == col]
    return _num(''.join(parts)) if parts else None


def _rmost(tokens):
    c = sorted([(x, v) for x, v in tokens if _band(x) == 'montant'])
    return _num(c[-1][1]) if c else None


def lire_facture(path):
    """Lit une facture PDF SOFRIPA -> list[LigneFacture]."""
    out = []
    with pdfplumber.open(path) as pdf:
        cur_exp = None
        for page in pdf.pages:
            lines = defaultdict(list)
            for w in page.extract_words(keep_blank_chars=False):
                lines[round(w['top'])].append((w['x0'], w['text']))
            cur = None
            for y in sorted(lines):
                toks = sorted(lines[y])
                txt = ' '.join(t for _, t in toks)
                m = re.search(r'Exp[ée]dition du (\d{2}/\d{2}/\d{2})', txt)
                if m:
                    cur_exp = m.group(1); continue
                if txt.startswith(('TOTAL', 'Report')) or 'Nbre OT' in txt:
                    cur = None; continue
                if 'EXP.' in txt:
                    cur = {'exp_date': cur_exp}; continue
                if 'DEST.:' in txt:
                    ot = [t for x, t in toks if _band(x) == 'date' and re.match(r'^\d{6,}$', t)]
                    d = [t for x, t in toks if _band(x) == 'desig']
                    if d and d[0] == 'DEST.:':
                        d = d[1:]
                    if cur is None:
                        cur = {'exp_date': cur_exp}
                    cur.update({'ot': ot[0] if ot else None, 'client': ' '.join(d), 'piece': None})
                    continue
                if cur is not None and re.match(r'^0000\d{4}$', txt.strip()):
                    cur['piece'] = txt.strip(); continue
                if cur is not None and 'ADMINISTRATIF' in txt:
                    out.append(cur); cur = None; continue
                if cur is not None and toks and toks[0][1] == 'FRAIS':
                    cur['poids'] = _merge(toks, 'poids')
                    mt = _rmost(toks)
                    if mt is not None:
                        cur['montant'] = mt
                    continue
                if cur is not None and ('KGS' in txt or 'PAL' in txt or txt.strip().startswith('...')):
                    cur['montant'] = _rmost(toks); continue
    # conversion dict -> LigneFacture
    return [LigneFacture(exp_date=x.get('exp_date'), ot=x.get('ot'), client=x.get('client'),
                         piece=x.get('piece'), poids=x.get('poids'), montant=x.get('montant'))
            for x in out]


def total_imprime_facture(path):
    """Somme des 'TOTAL JOURNALIER' imprimés sur la facture (contrôle de cohérence)."""
    txt = ""
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            txt += (pg.extract_text() or "") + "\n"
    printed = re.findall(r'TOTAL JOURNALIER DU \d{2}/\d{2}/\d{2}.*?([\d \u00a0]+,\d{2})\s*$', txt, re.M)
    return round(sum(_num(p) for p in printed), 2) if printed else None


def _find_col(headers, key):
    kl = key.lower()
    for i, h in enumerate(headers, 1):
        if str(h or '').strip().lower() == kl:
            return i
    for i, h in enumerate(headers, 1):
        if kl in str(h or '').lower():
            return i
    return None


def lire_commandes(path):
    """Lit l'export Easy Beer (Excel) -> dict[int, Commande]."""
    wb = load_workbook(path, data_only=True)
    ws = wb['Commandes'] if 'Commandes' in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ci = {k: _find_col(headers, v) for k, v in EB_KEYS.items()}
    if not ci['num']:
        raise ValueError("Colonne 'N° commande' introuvable dans l'export Easy Beer")

    commandes = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, ci['num']).value
        if not isinstance(n, (int, float)):
            continue
        brut = {headers[c - 1]: ws.cell(r, c).value for c in range(1, len(headers) + 1)}
        commandes[int(n)] = Commande(
            numero=int(n),
            client=ws.cell(r, ci['client']).value if ci['client'] else None,
            poids=ws.cell(r, ci['poids']).value if ci['poids'] else None,
            ht=ws.cell(r, ci['ht']).value if ci['ht'] else None,
            tournee=ws.cell(r, ci['tournee']).value if ci['tournee'] else None,
            brut=brut,
        )
    return commandes
