"""Tests de l'export XLSX palettes + de la requête list_sscc_for_ramasses.

Le builder est testé en relisant le fichier produit avec openpyxl.
La requête SQL est testée avec run_sql mocké (on vérifie le mapping + les
garde-fous, pas le SQL réel).
"""
from __future__ import annotations

import datetime as _dt
import io
from unittest.mock import MagicMock, patch

import openpyxl

from common.services.sscc_service import SsccLogEntry, list_sscc_for_ramasses
from common.xlsx_export import build_palettes_xlsx


def _entry(**kw) -> SsccLogEntry:
    base = dict(
        id=1,
        sscc="370000000000000017",
        user_email="op@x.fr",
        gtin_palette="03770000000017",
        lot="L240501",
        ddm=_dt.date(2026, 5, 1),
        case_count=84,
        generated_at=_dt.datetime(2026, 5, 1, 10, 30),
    )
    base.update(kw)
    return SsccLogEntry(**base)


# ─── build_palettes_xlsx ─────────────────────────────────────────────────────

class TestBuildPalettesXlsx:
    def _load(self, data: bytes):
        assert isinstance(data, bytes) and data[:2] == b"PK"  # zip magic
        wb = openpyxl.load_workbook(io.BytesIO(data))
        return wb.active

    def test_empty_has_header_only(self):
        ws = self._load(build_palettes_xlsx([]))
        assert ws.max_row == 1  # juste l'en-tête
        assert ws.cell(row=1, column=1).value == "Ramasse n°"
        assert ws.cell(row=1, column=4).value == "SSCC"

    def test_one_row_maps_fields(self):
        e = _entry(
            ramasse_numero=7,
            ramasse_date=_dt.date(2026, 5, 2),
            ramasse_destinataire="SOFRIPA",
            designation="Kéfir Gingembre",
            marque="NIKO",
            gout="Gingembre",
            loaded_at=_dt.datetime(2026, 5, 2, 8, 0),
        )
        ws = self._load(build_palettes_xlsx([e], sheet_title="Ramasses"))
        assert ws.title == "Ramasses"
        assert ws.max_row == 2
        # ligne de données (row 2)
        header = [c.value for c in ws[1]]
        row = [c.value for c in ws[2]]
        d = dict(zip(header, row))
        assert d["Ramasse n°"] == 7
        assert d["Date ramasse"] == "2026-05-02"
        assert d["Destinataire"] == "SOFRIPA"
        assert d["SSCC"] == "370000000000000017"
        assert d["Lot"] == "L240501"
        assert d["Produit"] == "Kéfir Gingembre"
        assert d["Cartons"] == 84
        assert d["DDM"] == "2026-05-01"
        assert d["Générée le"] == "2026-05-01 10:30"
        assert d["Chargée le"] == "2026-05-02 08:00"
        assert d["Annulée"] == "Non"
        assert d["Archivée"] == "Non"

    def test_voided_and_archived_flags(self):
        e = _entry(
            voided_at=_dt.datetime(2026, 5, 3, 9, 0),
            voided_reason="Doublon",
            label_archived_at=_dt.datetime(2026, 5, 3, 9, 5),
            label_archived_reason="Erreur",
        )
        ws = self._load(build_palettes_xlsx([e]))
        header = [c.value for c in ws[1]]
        d = dict(zip(header, [c.value for c in ws[2]]))
        assert d["Annulée"] == "Oui"
        assert d["Motif annulation"] == "Doublon"
        assert d["Archivée"] == "Oui"
        assert d["Motif archivage"] == "Erreur"

    def test_missing_ramasse_numero_blank(self):
        ws = self._load(build_palettes_xlsx([_entry(ramasse_numero=None)]))
        header = [c.value for c in ws[1]]
        d = dict(zip(header, [c.value for c in ws[2]]))
        # cellule vide → None à la relecture openpyxl
        assert d["Ramasse n°"] in (None, "")


# ─── list_sscc_for_ramasses ──────────────────────────────────────────────────

class TestListSsccForRamasses:
    @patch("common.services.sscc_service.run_sql")
    def test_empty_ids_skips_sql(self, m: MagicMock):
        assert list_sscc_for_ramasses("t1", []) == []
        m.assert_not_called()

    @patch("common.services.sscc_service.run_sql")
    def test_blank_ids_filtered_out(self, m: MagicMock):
        assert list_sscc_for_ramasses("t1", ["", "  "]) == []
        m.assert_not_called()

    @patch("common.services.sscc_service.run_sql")
    def test_maps_rows(self, m: MagicMock):
        m.return_value = [{
            "id": 5,
            "sscc": "370000000000000017",
            "user_email": "op@x.fr",
            "gtin_palette": "03770000000017",
            "lot": "L1",
            "ddm": _dt.date(2026, 5, 1),
            "case_count": 84,
            "generated_at": _dt.datetime(2026, 5, 1, 10, 0),
            "pl_ramasse_id": "r-1",
            "rh_numero": 3,
            "eph_designation": "Kéfir",
        }]
        out = list_sscc_for_ramasses("t1", ["r-1"])
        assert len(out) == 1
        assert out[0].ramasse_id == "r-1"
        assert out[0].ramasse_numero == 3
        assert out[0].designation == "Kéfir"
        # le param :rids est bien la liste nettoyée
        _, kwargs = m.call_args
        params = m.call_args[0][1]
        assert params["rids"] == ["r-1"]
        assert params["t"] == "t1"

    @patch("common.services.sscc_service.run_sql")
    def test_sql_error_returns_empty(self, m: MagicMock):
        m.side_effect = RuntimeError("boom")
        assert list_sscc_for_ramasses("t1", ["r-1"]) == []
